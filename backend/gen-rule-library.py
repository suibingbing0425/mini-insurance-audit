"""
读取 2025 版医保监管规则库 xlsx，生成 rule-library.json
- 第一个 sheet：88 条规则列表（序号/一级分类/二级分类/规则名称/是否对应知识点明细）
- 46 个明细 sheet：每条规则对应的知识点明细
"""
import openpyxl, json, os

src = r"E:\创智爱康实习文档\医疗保障基金智能监管规则库、知识库\2025年医疗保障基金智能监管规则库_手工整理版.xlsx"
out = os.path.join(os.path.dirname(__file__), "data", "rule-library.json")

print("正在读取 xlsx...")
wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

# 1. 读规则列表（第一个 sheet）
ws = wb["智能监管规则列表"]
rules = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    seq, cat1, cat2, name, has_detail = row[:5]
    if not name:
        continue
    rules.append({
        "seq": int(seq) if seq else len(rules) + 1,
        "category1": str(cat1 or "").strip(),
        "category2": str(cat2 or "").strip(),
        "name": str(name or "").strip(),
        "hasDetail": str(has_detail or "").strip() == "是",
        "knowledge": []
    })
print(f"规则列表：{len(rules)} 条")

# 2. 读知识点明细（按规则名匹配 sheet）
sheet_names = wb.sheetnames
for rule in rules:
    if not rule["hasDetail"]:
        continue
    # 匹配 sheet 名（规则名出现在 sheet 名里即匹配）
    matched = None
    for sn in sheet_names:
        if rule["name"] in sn:
            matched = sn
            break
    if not matched:
        continue
    ws2 = wb[matched]
    headers = None
    count = 0
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip() for c in row]
            continue
        if i > 500:  # 每条规则最多 500 行知识点（演示够用）
            break
        item = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                item[headers[j]] = str(val or "").strip() if val else ""
        if any(v for v in item.values()):
            rule["knowledge"].append(item)
            count += 1
    if count > 0:
        print(f"  {rule['name']}: {count} 条知识点")

# 3. 输出 JSON
os.makedirs(os.path.dirname(out), exist_ok=True)
with open(out, "w", encoding="utf-8") as f:
    json.dump({"rules": rules}, f, ensure_ascii=False, indent=2)

total_knowledge = sum(len(r["knowledge"]) for r in rules)
file_size = os.path.getsize(out) / 1024
print(f"\n完成：{len(rules)} 条规则，知识点共 {total_knowledge} 条")
print(f"文件：{out} ({file_size:.0f} KB)")
wb.close()
