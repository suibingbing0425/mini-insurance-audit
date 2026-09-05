import openpyxl

path = r"E:\创智爱康实习文档\医疗保障基金智能监管规则库、知识库\2025年医疗保障基金智能监管规则库_手工整理版.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

# 1. 完整规则列表
print("=== 智能监管规则列表（全部 88 条）===")
ws = wb["智能监管规则列表"]
for row in ws.iter_rows(values_only=True):
    print("  | ".join(str(c)[:25] if c is not None else "" for c in row))

# 2. 看几个关键明细 sheet 的前 6 行结构
for sheet in ["药品区分性别使用", "药品儿童禁用", "重复开药", "超说明书用量开药", "药品相互作用", "药品限支付疗程"]:
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n=== 「{sheet}」前 6 行（共 {ws.max_row} 行）===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6: break
            print("  " + " | ".join(str(c)[:30] if c is not None else "" for c in row))

wb.close()
